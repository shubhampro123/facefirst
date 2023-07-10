import openpyxl


def getRowCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def getColumnCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file, sheet_name, row_num, colum_no):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_num, column=colum_no).value


def writeData(file, sheet_name, row_num, column_no, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    # sheet.cell(row=row_num, column=column_no).value
    sheet.cell(row=row_num, column=column_no).value = data
    workbook.save(file)
